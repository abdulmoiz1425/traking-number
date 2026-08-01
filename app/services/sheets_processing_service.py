import os
import shutil
import time
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app.errors import FileValidationError
from app.services.column_utils import detect_net_payable_column, detect_status_column, detect_tracking_column
from app.services.google_auth_service import get_credentials
from app.services.google_sheets_service import (
    GREEN_BACKGROUND,
    RED_BACKGROUND,
    extract_spreadsheet_id,
    get_client,
    get_spreadsheet_title,
    highlight_rows,
    inspect_tab_columns,
    list_tabs,
    read_values,
    write_values,
)
from app.services.matching_engine import (
    find_status_matches,
    find_value_updates,
    read_tracking_numbers_with_status,
    read_tracking_numbers_with_value,
)
from app.services.session_utils import get_or_create_session_id, get_session_id

# session_id -> connected spreadsheet info / last result. In-memory,
# single-process — same pattern/limitation as the other _STORE dicts here.
_SHEETS_STORE = {}
_RESULT_STORE = {}


def cleanup_stale_output_folders(upload_folder, retention_minutes):
    if not os.path.isdir(upload_folder):
        return
    cutoff = time.time() - retention_minutes * 60
    for name in os.listdir(upload_folder):
        path = os.path.join(upload_folder, name)
        if not os.path.isdir(path):
            continue
        try:
            if os.path.getmtime(path) < cutoff:
                shutil.rmtree(path, ignore_errors=True)
        except OSError:
            continue


def connect_sheet(spreadsheet_id_or_url):
    if not spreadsheet_id_or_url or not spreadsheet_id_or_url.strip():
        raise FileValidationError("No Google Sheet was selected.")

    credentials = get_credentials()
    service = get_client(credentials)

    spreadsheet_id = extract_spreadsheet_id(spreadsheet_id_or_url)
    title = get_spreadsheet_title(service, spreadsheet_id)
    tabs = list_tabs(service, spreadsheet_id)
    if not tabs:
        raise FileValidationError("This Google Sheet has no worksheets.")

    session_id = get_or_create_session_id()
    _SHEETS_STORE[session_id] = {"spreadsheet_id": spreadsheet_id, "title": title}
    _RESULT_STORE.pop(session_id, None)

    tracking_tab = tabs[0]["title"]
    main_tab = tabs[1]["title"] if len(tabs) > 1 else tabs[0]["title"]
    net_payable_tab = tabs[2]["title"] if len(tabs) > 2 else tabs[0]["title"]

    tracking_info = _inspect_tab(service, spreadsheet_id, tracking_tab)
    main_info = _inspect_tab(service, spreadsheet_id, main_tab)
    net_payable_info = _inspect_tab(service, spreadsheet_id, net_payable_tab)

    return {
        "spreadsheet_title": title,
        "tabs": [tab["title"] for tab in tabs],
        "tracking_tab": tracking_tab,
        "main_tab": main_tab,
        "net_payable_tab": net_payable_tab,
        "tracking_file": tracking_info,
        "main_file": main_info,
        "net_payable_file": net_payable_info,
    }


def get_connected_sheet():
    session_id = get_session_id()
    store_entry = _SHEETS_STORE.get(session_id) if session_id else None
    if not store_entry:
        raise FileValidationError("No connected Google Sheet found for this session. Please select a sheet again.")
    return store_entry


def clear_sheets_session():
    session_id = get_session_id()
    if session_id:
        _SHEETS_STORE.pop(session_id, None)
        _RESULT_STORE.pop(session_id, None)


def set_last_result(result):
    session_id = get_or_create_session_id()
    _RESULT_STORE[session_id] = result


def get_last_result():
    session_id = get_session_id()
    return _RESULT_STORE.get(session_id) if session_id else None


def get_sheets_status():
    try:
        store_entry = get_connected_sheet()
        credentials = get_credentials()
    except FileValidationError:
        return {"connected": False}

    service = get_client(credentials)
    try:
        tabs = list_tabs(service, store_entry["spreadsheet_id"])
    except FileValidationError:
        return {"connected": False}

    status = {
        "connected": True,
        "spreadsheet_title": store_entry["title"],
        "tabs": [tab["title"] for tab in tabs],
    }

    last_result = get_last_result()
    if last_result:
        status["result"] = last_result

    return status


def get_columns_for_tab(tab_title):
    store_entry = get_connected_sheet()
    credentials = get_credentials()
    service = get_client(credentials)
    return _inspect_tab(service, store_entry["spreadsheet_id"], tab_title)


def process_sheets(
    tracking_tab,
    tracking_column_letter,
    tracking_status_column_letter,
    tracking_has_header,
    main_tab,
    main_column_letter,
    main_has_header,
    net_payable_tab,
    net_payable_tracking_column_letter,
    net_payable_value_column_letter,
    net_payable_has_header,
    upload_folder,
):
    store_entry = get_connected_sheet()
    credentials = get_credentials()
    service = get_client(credentials)
    spreadsheet_id = store_entry["spreadsheet_id"]

    # Session-scoped so concurrent users never collide on the fixed
    # "unmatched_tracking_numbers.xlsx" filename, and so a stale session's
    # directory (this file included) is what cleanup_stale_output_folders
    # actually removes.
    session_id = get_or_create_session_id()
    output_folder = os.path.join(upload_folder, session_id, "output")

    tracking_col_idx = column_index_from_string(tracking_column_letter)
    status_col_idx = column_index_from_string(tracking_status_column_letter)
    tracking_values = read_values(service, spreadsheet_id, tracking_tab)
    tracking_result = read_tracking_numbers_with_status(
        tracking_values, tracking_col_idx, status_col_idx, tracking_has_header
    )

    status_map = tracking_result["status_map"]
    if not status_map:
        raise FileValidationError("No valid tracking numbers were found in the tracking-number tab.")

    main_col_idx = column_index_from_string(main_column_letter)
    main_values = read_values(service, spreadsheet_id, main_tab)

    min_data_row = 2 if main_has_header else 1
    if len(main_values) < min_data_row:
        raise FileValidationError("The selected main tab contains no data.")

    match_result = find_status_matches(main_values, main_col_idx, status_map, main_has_header)
    if match_result["non_blank_scanned"] == 0:
        raise FileValidationError("The selected main-tab tracking-number column contains no values.")

    tabs = list_tabs(service, spreadsheet_id)
    main_tab_info = next((tab for tab in tabs if tab["title"] == main_tab), None)
    if main_tab_info is None:
        raise FileValidationError("The selected tab could not be found. Please reconnect the sheet.")

    row_colors = {}
    for row_number in match_result["delivered_row_numbers"]:
        row_colors[row_number] = GREEN_BACKGROUND
    for row_number in match_result["return_row_numbers"]:
        row_colors[row_number] = RED_BACKGROUND

    total_columns = max((len(row) for row in main_values), default=0)
    highlight_rows(service, spreadsheet_id, main_tab_info["sheet_id"], row_colors, total_columns)

    matched_set = match_result["matched_values"]
    unmatched_set = set(status_map.keys()) - matched_set

    np_tracking_col_idx = column_index_from_string(net_payable_tracking_column_letter)
    np_value_col_idx = column_index_from_string(net_payable_value_column_letter)
    net_payable_values = read_values(service, spreadsheet_id, net_payable_tab)
    net_payable_result = read_tracking_numbers_with_value(
        net_payable_values, np_tracking_col_idx, np_value_col_idx, net_payable_has_header
    )
    value_map = net_payable_result["value_map"]
    if not value_map:
        raise FileValidationError("No valid tracking numbers were found in the Net Payable tab.")

    main_columns = inspect_tab_columns(main_values)
    detected_np_letter = detect_net_payable_column(main_columns)
    if detected_np_letter:
        net_payable_main_col_idx = column_index_from_string(detected_np_letter)
        write_header = False
    else:
        # No existing "Net Payable" column in the main sheet - append a new
        # one right after the last used column instead of overwriting data.
        net_payable_main_col_idx = total_columns + 1
        write_header = main_has_header

    value_update_result = find_value_updates(main_values, main_col_idx, value_map, main_has_header)
    cell_updates = {
        (row_number, net_payable_main_col_idx): value for row_number, value in value_update_result["updates"].items()
    }
    if write_header:
        cell_updates[(1, net_payable_main_col_idx)] = "Net Payable"

    write_values(service, spreadsheet_id, main_tab, cell_updates)

    os.makedirs(output_folder, exist_ok=True)
    timestamp = datetime.now()
    unmatched_filename = "unmatched_tracking_numbers.xlsx"
    unmatched_path = os.path.join(output_folder, unmatched_filename)
    _write_unmatched_workbook(unmatched_set, unmatched_path)

    summary = {
        "total_tracking_numbers_read": tracking_result["total_read"],
        "blank_tracking_cells_ignored": tracking_result["blanks_ignored"],
        "duplicate_tracking_numbers_removed": tracking_result["duplicates_removed"],
        "unique_tracking_numbers_searched": len(status_map),
        "tracking_numbers_matched": len(matched_set),
        "tracking_numbers_not_matched": len(unmatched_set),
        "rows_marked_delivered": len(match_result["delivered_row_numbers"]),
        "rows_marked_return": len(match_result["return_row_numbers"]),
        "rows_with_unrecognized_status": len(match_result["unrecognized_status_rows"]),
        "total_rows_highlighted": len(row_colors),
        "net_payable_rows_updated": len(value_update_result["updates"]),
        "net_payable_column": get_column_letter(net_payable_main_col_idx),
        "processing_status": "success",
        "processing_datetime": timestamp.isoformat(timespec="seconds"),
    }

    return {
        "summary": summary,
        "unmatched_filename": unmatched_filename,
        "unmatched_path": unmatched_path,
        "main_sheet_url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit",
    }


def _inspect_tab(service, spreadsheet_id, tab_title):
    values = read_values(service, spreadsheet_id, tab_title)
    if not values:
        raise FileValidationError(f'The tab "{tab_title}" contains no data.')

    columns = inspect_tab_columns(values)
    detected_column = detect_tracking_column(columns)
    detected_status_column = detect_status_column(columns)
    detected_net_payable_column = detect_net_payable_column(columns)

    return {
        "columns": columns,
        "detected_column": detected_column,
        "detected_status_column": detected_status_column,
        "detected_net_payable_column": detected_net_payable_column,
    }


def _write_unmatched_workbook(unmatched_set, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Unmatched"
    ws.append(["Tracking Number"])
    for value in sorted(unmatched_set):
        ws.append([value])
    wb.save(path)
